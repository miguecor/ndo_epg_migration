#!/usr/bin/env python3

import argparse
import json
import logging
import os.path
import sys
import requests
import shutil

from credentials import *
from datetime import datetime
from logging import Logger
from openpyxl import load_workbook
from os import makedirs, path
from requests import Session, Response
from requests.exceptions import RequestException
from time import sleep
from tqdm import tqdm, trange
from urllib3 import disable_warnings


disable_warnings()

# Argument Parser
parser = argparse.ArgumentParser(description="Reads/writes an Excel file to migrate EPGs between 2 tenants in an NDO")
parser.add_argument("-f", "--filename", type=str, nargs="?", default="data.xlsx",
                    help="Name of the file with src and dst EPG information.  If a filename is not provided, will "
                         "use 'data.xlsx'")
group = parser.add_mutually_exclusive_group(required=True)
group.add_argument("-g", "--get", action="store_true",
                    help="Connects to the NSO to obtain all the information necessary to create an Excel file for "
                         "the EPG migration.  The name of the output file can be specified with '--filename', otherwise"
                         " 'data.xlsx'")
group.add_argument("-p", "--put", action="store_true",
                    help="Uses the provided file to migrate EPGs based on the data in the spreadsheet. The name "
                         "of the input file can be specified with '--filename', otherwise 'data.xlsx'")
parser.add_argument("-s", "--ssl", action="store_true",
                    help="Enables the NDO certificate validation, default is no validation")
parser.add_argument("-d", "--debug", action="store_true",
                    help="Enables debugging on the terminal for a more verbose output.  Default console output is INFO")
parser = parser.parse_args()

# Logging Parameters
TIMESTAMP = datetime.now().strftime("%Y%m%d")
LOGGER = "ndo_epg_migration"
LOG_FOLDER = ".logs"
LOG_FILE = "ndo_epg_migration_%s.log" % TIMESTAMP
LOG_FORMAT = "%(asctime)s - %(levelname)s - %(message)s"
if not path.exists(LOG_FOLDER):
    makedirs(LOG_FOLDER)


# Logger Definition
def get_logger() -> Logger:
    """Local file and console logging"""
    logger = logging.getLogger(LOGGER)
    logger.setLevel(logging.DEBUG)
    formatter = logging.Formatter(LOG_FORMAT)
    # File handler
    fh = logging.FileHandler((".logs/%s" % LOG_FILE))
    fh.setFormatter(formatter)
    fh.setLevel(logging.DEBUG)
    # Console handler
    ch = logging.StreamHandler()
    if parser.debug:
        ch.setLevel(logging.DEBUG)
    else:
        ch.setLevel(logging.INFO)
    ch.setFormatter(formatter)
    # Setup logger
    logger.addHandler(fh)
    logger.addHandler(ch)
    return logger


log = get_logger()


# Decorators:
def request_logger(func):
    def wrapper(*args, **kwargs):
        try:
            content = func(*args, **kwargs)
            if content is not None and type(content) == Session:
                log.debug(str(content.cookies)[:80])
                return content
            elif content is not None and type(content) == Response:
                log.debug("%(method)s: %(url)s STATUS: %(status)s" % dict(
                    url=content.request.url, method=content.request.method, status=content.status_code))
                return content
        except RequestException as e:
            log.critical(e)
    return wrapper


# Function Definitions:
@request_logger
def get_ndo_session(**kwargs) -> Session:
    url = "https://%s" % NDO_IP
    api = "/login"
    payload = {"domain": NDO_DOM, "username": NDO_USR, "userPasswd": NDO_PWD}
    headers = {'Content-Type': 'application/json'}
    session = requests.session()
    response = session.post(url + api, headers=headers, data=json.dumps(payload), **kwargs)
    response.raise_for_status()
    log.info("User '%(user)s' successfully authenticated to %(url)s" % dict(user=payload['username'], url=url))
    return session


@request_logger
def get_ndo_sites(session: Session, **kwargs):
    url = "https://%s" % NDO_IP
    api = "/mso/api/v2/sites"
    response = session.get(url + api, **kwargs)
    response.raise_for_status()
    log.info("Received all sites information from %s" % url)
    return response


@request_logger
def get_ndo_tenants(session: Session, **kwargs):
    url = "https://%s" % NDO_IP
    api = "/mso/api/v1/tenants"
    response = session.get(url + api, **kwargs)
    response.raise_for_status()
    log.info("Received all tenants information from %s" % url)
    return response


@request_logger
def get_ndo_schemas(session: Session, **kwargs):
    url = "https://%s" % NDO_IP
    api = "/mso/api/v1/schemas"
    response = session.get(url + api, **kwargs)
    response.raise_for_status()
    log.info("Received all schemas information from %s" % url)
    return response


@request_logger
def patch_ndo_tmpl_bds(session: Session, data: dict, bdRef="", oper="", **kwargs):
    allowed_ops = ["add", "remove"]
    if oper not in allowed_ops:
        raise ValueError("Unsupported PATCH operation")
    schema_id = bdRef.split("/")[2]
    templ_name = bdRef.split("/")[4]
    url = "https://%s" % NDO_IP
    api = "/mso/api/v1/schemas/%s" % schema_id
    payload = []

    if oper == "add":
        payload = [{"op": "add",
                    "path": "/templates/%s/bds/-" % templ_name,
                    "value": data}]
    elif oper == "remove":
        payload = [{"op": "remove",
                    "path": "/templates/%s/bds/%s" % (templ_name, data['name'])}]

    response = session.patch(url + api, data=json.dumps(payload), **kwargs)
    response.raise_for_status()
    log.info("Template BD patch %s operation returned status code %s" % (oper, response.status_code))
    return response


@request_logger
def patch_ndo_site_bds(session: Session, data: dict, siteId="", oper="", **kwargs):
    allowed_ops = ["remove", "replace"]
    if oper not in allowed_ops:
        raise ValueError("Unsupported PATCH operation")
    schema_id = data['bdRef'].split("/")[2]
    templ_name = data['bdRef'].split("/")[4]
    bd_name = data['bdRef'].split("/")[-1]
    url = "https://%s" % NDO_IP
    api = "/mso/api/v1/schemas/%s" % schema_id
    payload = []
    if oper == "remove":
        payload = [{"op": oper,
                    "path": "/sites/%s-%s/bds/%s" % (siteId, templ_name, bd_name)}]
    elif oper == "replace":
        payload = [{"op": oper,
                    "path": "/sites/%s-%s/bds/%s" % (siteId, templ_name, bd_name),
                    "value": data}]

    response = session.patch(url + api, data=json.dumps(payload), **kwargs)
    response.raise_for_status()
    log.info("Site BD patch %s operation returned status code %s" % (oper, response.status_code))
    return response


@request_logger
def patch_ndo_tmpl_epgs(session: Session, data: dict, epgRef="", oper="", **kwargs):
    allowed_ops = ["add", "remove"]
    if oper not in allowed_ops:
        raise ValueError("Unsupported PATCH operation")
    schema_id = epgRef.split("/")[2]
    templ_name = epgRef.split("/")[4]
    anp_name = epgRef.split("/")[6]
    url = "https://%s" % NDO_IP
    api = "/mso/api/v1/schemas/%s" % schema_id
    payload = []

    if oper == "add":
        payload = [{"op": "add",
                    "path": "/templates/%s/anps/%s/epgs/-" % (templ_name, anp_name),
                    "value": data}]
    elif oper == "remove":
        payload = [{"op": "remove",
                    "path": "/templates/%s/anps/%s/epgs/%s" % (templ_name, anp_name, data['name'])}]

    response = session.patch(url + api, data=json.dumps(payload), **kwargs)
    response.raise_for_status()
    log.info("Template EPG patch %s operation returned status code %s" % (oper, response.status_code))
    return response


@request_logger
def patch_ndo_site_epgs(session: Session, data: dict, siteId="", oper="", **kwargs):
    allowed_ops = ["remove", "replace"]
    if oper not in allowed_ops:
        raise ValueError("Unsupported PATCH operation")
    schema_id = data['epgRef'].split("/")[2]
    templ_name = data['epgRef'].split("/")[4]
    anp_name = data['epgRef'].split("/")[6]
    epg_name = data['epgRef'].split("/")[-1]
    url = "https://%s" % NDO_IP
    api = "/mso/api/v1/schemas/%s" % schema_id
    payload = []

    if oper == "remove":
        payload = [{"op": oper,
                    "path": "/sites/%s-%s/anps/%s/epgs/%s" % (siteId, templ_name, anp_name, epg_name)}]
    elif oper == "replace":
        payload = [{"op": oper,
                    "path": "/sites/%s-%s/anps/%s/epgs/%s" % (siteId, templ_name, anp_name, epg_name),
                    "value": data}]

    response = session.patch(url + api, data=json.dumps(payload), **kwargs)
    response.raise_for_status()
    log.info("Site EPG patch %s operation returned status code %s" % (oper, response.status_code))
    return response


@request_logger
def patch_ndo_epg_static_ports(session: Session, data, siteId="", ports=[], oper="", **kwargs):
    allowed_ops = ["remove", "replace"]
    if oper not in allowed_ops:
        raise ValueError("Unsupported PATCH operation")
    schema_id = data['epgRef'].split("/")[2]
    templ_name = data['epgRef'].split("/")[4]
    anp_name = data['epgRef'].split("/")[6]
    epg_name = data['epgRef'].split("/")[-1]
    url = "https://%s" % NDO_IP
    api = "/mso/api/v1/schemas/%s" % schema_id
    if oper == "remove":
        payload = [{"op": oper,
                    "path": "/sites/%s-%s/anps/%s/epgs/%s/staticPorts" % (siteId, templ_name, anp_name, epg_name)}]
    elif oper == "replace":
        payload = [{"op": oper,
                    "path": "/sites/%s-%s/anps/%s/epgs/%s/staticPorts" % (siteId, templ_name, anp_name, epg_name),
                    "value": ports}]

    response = session.patch(url + api, data=json.dumps(payload), **kwargs)
    response.raise_for_status()
    log.info("EPG static ports patch %s operation returned status code %s" % (oper, response.status_code))
    return response


@request_logger
def deploy_ndo_template(session: Session, schm="", tmpl="", **kwargs):
    url = "https://%s" % NDO_IP
    api = "/mso/api/v1/task"
    payload = {"isRedeploy": False, "schemaId": schm, "templateName": tmpl}
    response = session.post(url + api, data=json.dumps(payload), **kwargs)
    response.raise_for_status()
    log.debug("%s" % response.json())
    log.info("Deployment task status: %s" % response.status_code)
    return response


@request_logger
def ndo_deploy_status_check(session: Session, id="", **kwargs):
    url = "https://%s" % NDO_IP
    api = "/mso/api/v1/deployments/%s" % id
    task_status = ""
    response = Response
    while task_status != "Complete":
        response = session.get(url + api, **kwargs)
        response.raise_for_status()
        task_status = response.json()['operDetails']['taskStatus']
        if task_status == "Error":
            log.error(response.json()['operDetails']['execSiteStatus'][0]['status']['msg'][:-3])
            raise RequestException("Deployment task %s failed" % id)
        log.info("Deployment task ID %s status is %s" % (id, task_status))
        sleep(0.5)
    return response


def write_to_excel(filename: str, sheetname: str, data: list, startCol=1, startRow=1):
    if not os.path.exists(filename):
        template = "template.xlsx"
        shutil.copy(template, filename)
    log.info("Successfully added data to file %s from template!" % filename)
    workbook = load_workbook(filename)
    # sheet = workbook.active
    sheet = workbook[sheetname]

    column_names = list(data[0].keys())
    for col_index, column_name in enumerate(column_names, start=startCol):
        sheet.cell(row=startRow, column=col_index).value = column_name

    for row_index, row_data in enumerate(data, start=(startRow + 1)):
        for col_index, column_name in enumerate(column_names, start=startCol):
            sheet.cell(row=row_index, column=col_index).value = row_data[column_name]

    workbook.save(filename)
    log.info(f"Data written to {filename} successfully!")


def normalize_sites_data(sites_data: dict):
    data = []
    for site in sites_data['sites']:
        commonId = "None"
        siteId = site['id'] or "None"
        commonName = site['common']['name'] or "None"
        commonDisplayName = site['common']['displayName'] or "None"
        try:
            commonId = site['common']['siteId'] or "None"
        except KeyError as e:
            log.debug("Site %(name)s noes not have %(key)s, return None" % dict(name=commonName, key=e))
        log.debug("Site Data: Name %(name)s, Id %(id)s, No %(site)s" % dict(
            name=commonName, site=commonId, id=siteId))
        data.append(
            {'Site Name': commonName, 'Site Display Name': commonDisplayName, 'Site ID': siteId,
             'Site Number': commonId}
        )
        log.info("Normalized %s site data" % commonName)
    return data


def normalize_tenants_data(tenants_data: dict):
    data = []
    for tenant in tenants_data['tenants']:
        tnId = tenant['id'] or "None"
        tnName = tenant['name'] or "None"
        tnDisplayName = tenant['displayName'] or "None"
        for association in tenant['siteAssociations']:
            siteAssociation = association['siteId']
            data.append(
                {'Tenant Name': tnName, 'Tenant Display Name': tnDisplayName, 'Tenand ID': tnId,
                 'Site Association': siteAssociation}
            )
        log.info("Normalized %s tenant data" % tnDisplayName)
    return data


def normalize_schema_template_data(schemas_data: dict):
    data = []
    for schema in schemas_data['schemas']:
        schmDisplayName = schema['displayName']
        schmId = schema['id']
        schmVersion = schema['_updateVersion']
        for tmplNum, template in enumerate(schema['templates']):
            tmplName = template['name']
            tmplDisplayName = template['displayName']
            tmplID = template['templateID']
            tmplVersion = template['version']
            tmplTnId = template['tenantId']
            tmplType = template['templateType']
            data.append(
                {'Schema Display Name': schmDisplayName, 'Schema ID': schmId, 'Schema Version': schmVersion,
                 'Template Name': tmplName, 'Template Display Name': tmplDisplayName, 'Template Number': tmplNum,
                 'Template ID': tmplID, 'Template Type': tmplType, 'Template Version': tmplVersion,
                 'Associated Tenant ID': tmplTnId}
            )
            log.debug("Normalized %(schm)s schema %(tmpl)s template data" % dict(
                schm=schmDisplayName, tmpl=tmplDisplayName)
                     )
        log.info("Normalized %(schm)s schema template data" % dict(schm=schmDisplayName))
    return data


def normalize_schema_site_data(schemas_data: dict):
    data = []
    for schema in schemas_data['schemas']:
        schmDisplayName = schema['displayName']
        schmId = schema['id']
        schmVersion = schema['_updateVersion']
        for site in schema['sites']:
            siteId, siteTmplName, siteTmplID = ("None",) * 3
            if len(site) >= 1:
                siteId = site['siteId']
                siteTmplName = site['templateName']
                siteTmplID = site['templateID']
            elif len(site) == 0:
                log.warning("Schema %s is not associated to any site yet" % schmDisplayName)
            data.append(
                {'Schema Display Name': schmDisplayName, 'Schema ID': schmId, 'Schema Version': schmVersion,
                 'Site ID': siteId, 'Site Template Name': siteTmplName, 'Site Template ID': siteTmplID}
            )
            log.debug("Normalized %(schm)s schema %(site)s site data" % dict(schm=schmDisplayName, site=siteId))
        log.info("Normalized %(schm)s schema site data" % dict(schm=schmDisplayName))
    return data


def normalize_vrf_template_data(schemas_data: dict):
    data = []
    for schema in schemas_data['schemas']:
        schmId = schema['id']
        schmDisplayName = schema['displayName']
        dict1 = {'Schema ID': schmId}
        for tmplNum, template in enumerate(schema['templates']):
            tmplName = template['name']
            tmplID = template['templateID']
            tmplTnId = template['tenantId']
            dict2 = {'Template ID': tmplID, 'Tenant ID': tmplTnId}
            log.debug(json.dumps({**dict1, **dict2}))
            vrfs = template['vrfs']
            for vrfNum, vrf in enumerate(vrfs):
                vrfName, vrfDisplayName, vrfUUID, vrfRef, dict3 = ("None",) * 5
                if len(vrfs) >= 1:
                    vrfName = vrf['name']
                    vrfDisplayName = vrf['displayName']
                    vrfUUID = vrf['uuid']
                    vrfRef = vrf['vrfRef']
                    dict3 = {'VRF Name': vrfName, 'VRF Display Name': vrfDisplayName, 'VRF Number': vrfNum,
                             'VRF UUID': vrfUUID, 'VRF Reference': vrfRef}
                    log.debug(json.dumps(dict3))
                elif len(vrfs) == 0:
                    vrfNum = "None"
                    log.warning("/%(schm)s/%(tmpl)s template does not have any VRF created" % dict(
                        schm=schmDisplayName, tmpl=tmplName))
                    dict3 = {'VRF Name': vrfName, 'VRF Display Name': vrfDisplayName, 'VRF Number': vrfNum,
                             'VRF UUID': vrfUUID, 'VRF Reference': vrfRef}
                    log.debug(json.dumps(dict3))
                vrf_dict = {**dict1, **dict2, **dict3}
                data.append(vrf_dict)
                log.info("Normalized /%(schm)s/%(tmpl)s/%(vrf)s VRF template data" % dict(
                    schm=schmDisplayName, tmpl=tmplName, vrf=vrfName
                ))
    return data


def normalize_vrf_site_data(schemas_data: dict):
    data = []
    for schema in schemas_data['schemas']:
        schmId = schema['id']
        schmDisplayName = schema['displayName']
        dict1 = {'Schema ID': schmId}
        if len(schema['sites']) >= 1:
            for siteNum, site in enumerate(schema['sites']):
                siteId = site['siteId']
                siteTmplName = site['templateName']
                siteTmplID = site['templateID']
                dict2 = {'Site ID': siteId, 'Site Template ID': siteTmplID}
                log.debug(json.dumps({**dict1, **dict2}))
                vrfs = site['vrfs']
                for vrfNum, vrf in enumerate(vrfs):
                    vrfRef, dict3 = ("None",) * 2
                    if len(vrfs) >= 1:
                        vrfRef = vrf['vrfRef']
                    elif len(vrfs) == 0:
                        vrfNum = "None"
                        log.warning(
                            "/%(schm)s/%(tmpl)s template does not have any VRF at the site (%(site)s) level" % dict(
                                schm=schmDisplayName, tmpl=siteTmplName, site=siteId))
                    dict3 = {'VRF Number': vrfNum, 'VRF Reference': vrfRef}
                    log.debug(json.dumps(dict3))
                    vrf_dict = {**dict1, **dict2, **dict3}
                    data.append(vrf_dict)
                    log.info("Normalized /%(schm)s/%(tmpl)s/%(vrf)s VRF site data" % dict(
                        schm=schmDisplayName, tmpl=siteTmplName, vrf=vrfRef
                    ))
        elif len(schema['sites']) == 0:
            log.warning("%s schema is not currently associated to any site" % schmDisplayName)
    return data


def normalize_bd_template_data(schemas_data: dict):
    data = []
    for schema in schemas_data['schemas']:
        schmId = schema['id']
        schmDisplayName = schema['displayName']
        dict1 = {'Schema ID': schmId}
        for tmplNum, template in enumerate(schema['templates']):
            tmplName = template['name']
            tmplID = template['templateID']
            tmplTnId = template['tenantId']
            dict2 = {'Template ID': tmplID, 'Tenant ID': tmplTnId}
            log.debug(json.dumps({**dict1, **dict2}))
            bds = template['bds']
            for bdNum, bd in enumerate(bds):
                bdName, bdDisplayName, bdUUID, bdRef, bdL2Stretch, bdSubnets, bdVrfRef, dict3 = ("None",) * 8
                if len(bds) >= 1:
                    bdName = bd['name']
                    bdDisplayName = bd['displayName']
                    bdUUID = bd['uuid']
                    bdRef = bd['bdRef']
                    bdL2Stretch = bd['l2Stretch']
                    subnets = bd['subnets']
                    if subnets:
                        bdSubnets = True
                    elif not subnets:
                        bdSubnets = False
                    bdVrfRef = bd['vrfRef']
                    log.debug(json.dumps(dict3))
                elif len(bds) == 0:
                    bdNum = "None"
                    log.warning("/%(schm)s/%(tmpl)s template does not have any BD created" % dict(
                        schm=schmDisplayName, tmpl=tmplName))
                dict3 = {'BD Name': bdName, 'BD Display Name': bdDisplayName, 'BD Number': bdNum,
                         'BD UUID': bdUUID, 'BD Reference': bdRef, 'L2 Stretch': bdL2Stretch,
                         'BD Template Subnets': bdSubnets, 'BD VRF Reference': bdVrfRef}
                log.debug(json.dumps(dict3))
                bd_dict = {**dict1, **dict2, **dict3}
                data.append(bd_dict)
                log.info("Normalized /%(schm)s/%(tmpl)s/%(bd)s BD template data" % dict(
                    schm=schmDisplayName, tmpl=tmplName, bd=bdName
                ))
    return data


def normalize_bd_site_data(schemas_data: dict):
    data = []
    for schema in schemas_data['schemas']:
        schmId = schema['id']
        schmDisplayName = schema['displayName']
        dict1 = {'Schema ID': schmId}
        if len(schema['sites']) >= 1:
            for siteNum, site in enumerate(schema['sites']):
                siteId = site['siteId']
                siteTmplName = site['templateName']
                siteTmplID = site['templateID']
                dict2 = {'Site ID': siteId, 'Site Template ID': siteTmplID}
                log.debug(json.dumps({**dict1, **dict2}))
                bds = site['bds']
                for bdNum, bd in enumerate(bds):
                    bdRef, bdSubnets, dict3 = ("None",) * 3
                    if len(bds) >= 1:
                        bdRef = bd['bdRef']
                        subnets = bd['subnets']
                        if subnets:
                            bdSubnets = True
                        elif not subnets:
                            bdSubnets = False
                    elif len(bds) == 0:
                        bdNum = "None"
                        log.warning(
                            "/%(schm)s/%(tmpl)s template does not have any BD at the site (%(site)s) level" % dict(
                                schm=schmDisplayName, tmpl=siteTmplName, site=siteId))
                    dict3 = {'BD Number': bdNum, 'BD Reference': bdRef, 'BD Site Subnets': bdSubnets}
                    log.debug(json.dumps(dict3))
                    bd_dict = {**dict1, **dict2, **dict3}
                    data.append(bd_dict)
                    log.info("Normalized /%(schm)s/%(tmpl)s/%(bd)s BD site data" % dict(
                        schm=schmDisplayName, tmpl=siteTmplName, bd=bdRef
                    ))
        elif len(schema['sites']) == 0:
            log.warning("%s schema is not currently associated to any site" % schmDisplayName)
    return data


def normalize_epg_template_data(schemas_data: dict):
    data = []
    for schema in schemas_data['schemas']:
        schmId = schema['id']
        schmDisplayName = schema['displayName']
        dict1 = {'Schema ID': schmId}
        for tmplNum, template in enumerate(schema['templates']):
            tmplName = template['name']
            tmplID = template['templateID']
            tmplTnId = template['tenantId']
            dict2 = {'Template ID': tmplID, 'Tenant ID': tmplTnId}
            log.debug(json.dumps({**dict1, **dict2}))
            anps = template['anps']
            for anpNum, anp in enumerate(anps):
                anpName, anpDisplayName, anpUUID, anpRef = ("None",) * 4
                epgName, epgDisplayName, epgUUID, epgRef, epgBdRef = ("None",) * 5
                if len(anps) >= 1:
                    anpName = anp['name']
                    anpDisplayName = anp['displayName']
                    anpUUID = anp['uuid']
                    anpRef = anp['anpRef']
                    dict3 = {'ANP Name': anpName, 'ANP Display Name': anpDisplayName, 'ANP Number': anpNum,
                             'ANP UUID': anpUUID, 'ANP Reference': anpRef}
                    log.debug(json.dumps(dict3))
                    epgs = anp['epgs']
                    for epgNum, epg in enumerate(epgs):
                        if len(epgs) >= 1:
                            epgName = epg['name']
                            epgDisplayName = epg['displayName']
                            epgUUID = epg['uuid']
                            epgRef = epg['epgRef']
                            epgBdRef = epg['bdRef']
                        elif len(epgs) == 0:
                            epgNum = "None"
                            log.warning("Application Profile %(anp)s does not have any EPGs created" % dict(
                                anp=anpName))
                        dict4 = {'EPG Name': epgName, 'EPG Display Name': epgDisplayName, 'EPG Number': epgNum,
                                 'EPG UUID': epgUUID, 'EPG Reference': epgRef, 'EPG BD Reference': epgBdRef}
                        log.debug(json.dumps(dict4))
                        epg_dict = {**dict1, **dict2, **dict3, **dict4}
                        data.append(epg_dict)
                elif len(anps) == 0:
                    anpNum, epgNum = ("None",) * 2
                    log.warning("/%(schm)s/%(tmpl)s template does not have any Application Profiles created" % dict(
                        schm=schmDisplayName, tmpl=tmplName))
                    dict3 = {'ANP Name': anpName, 'ANP Display Name': anpDisplayName, 'ANP Number': anpNum,
                             'ANP UUID': anpUUID, 'ANP Reference': anpRef}
                    dict4 = {'EPG Name': epgName, 'EPG Display Name': epgDisplayName, 'EPG Number': epgNum,
                             'EPG UUID': epgUUID, 'EPG Reference': epgRef, 'EPG BD Reference': epgBdRef}
                    log.debug(json.dumps(dict3))
                    log.debug(json.dumps(dict4))
                    epg_dict = {**dict1, **dict2, **dict3, **dict4}
                    data.append(epg_dict)
                log.info("Normalized /%(schm)s/%(tmpl)s/%(anp)s app profile template data" % dict(
                    schm=schmDisplayName, tmpl=tmplName, anp=anpDisplayName
                ))
    return data


def normalize_epg_site_data(schemas_data: dict):
    data = []
    for schema in schemas_data['schemas']:
        anpRef, epgRef, siteId = ("None",) * 3
        schmId = schema['id']
        schmDisplayName = schema['displayName']
        dict1 = {'Schema ID': schmId}
        if len(schema['sites']) >= 1:
            for siteNum, site in enumerate(schema['sites']):
                siteId = site['siteId']
                siteTmplName = site['templateName']
                siteTmplID = site['templateID']
                dict2 = {'Site ID': siteId, 'Site Template ID': siteTmplID}
                log.debug(json.dumps({**dict1, **dict2}))
                anps = site['anps']
                for anpNum, anp in enumerate(anps):
                    if len(anps) >= 1:
                        anpRef = anp['anpRef']
                        dict3 = {'ANP Number': anpNum, 'ANP Reference': anpRef}
                        epgs = anp['epgs']
                        for epgNum, epg in enumerate(epgs):
                            if len(epgs) >= 1:
                                epgRef = epg['epgRef']
                            elif len(epgs) == 0:
                                epgNum = "None"
                                log.warning(
                                    "App profile %(anp)s does not have any EPGs created at the site level %(site)s" % dict(
                                        anp=anpName, site=siteId))
                            dict4 = {'EPG Number': epgNum, 'EPG Reference': epgRef}
                            log.debug(json.dumps({**dict3, **dict4}))
                            epg_dict = {**dict1, **dict2, **dict3, **dict4}
                            data.append(epg_dict)
                    elif len(anps) == 0:
                        anpNum, epgNum = ("None",) * 2
                        log.warning(
                            "/%(schm)s/%(tmpl)s template does not have any app profile at the site (%(site)s) level" % dict(
                                schm=schmDisplayName, tmpl=siteTmplName, site=siteId))
                        dict3 = {'ANP Number': anpNum, 'ANP Reference': anpRef}
                        dict4 = {'EPG Number': epgNum, 'EPG Reference': epgRef}
                        log.debug(json.dumps({**dict3, **dict4}))
                        epg_dict = {**dict1, **dict2, **dict3, **dict4}
                        data.append(epg_dict)
                    log.info("Normalized /%(schm)s/%(tmpl)s/%(anp)s app profile site data" % dict(
                        schm=schmDisplayName, tmpl=siteTmplName, anp=anpRef
                    ))
        elif len(schema['sites']) == 0:
            log.warning("%s schema is not currently associated to any site" % schmDisplayName)
    return data


def normalize_contract_template_data(schemas_data: dict):
    data = []
    for schema in schemas_data['schemas']:
        schmId = schema['id']
        schmDisplayName = schema['displayName']
        dict1 = {'Schema ID': schmId}
        for tmplNum, template in enumerate(schema['templates']):
            tmplName = template['name']
            tmplID = template['templateID']
            tmplTnId = template['tenantId']
            dict2 = {'Template ID': tmplID, 'Tenant ID': tmplTnId}
            log.debug(json.dumps({**dict1, **dict2}))
            contracts = template['contracts']
            for cntrNum, cntr in enumerate(contracts):
                cntrName, cntrDisplayName, cntrUUID, cntrRef, dict3 = ("None",) * 5
                if len(contracts) >= 1:
                    cntrName = cntr['name']
                    cntrDisplayName = cntr['displayName']
                    cntrUUID = cntr['uuid']
                    cntrRef = cntr['contractRef']
                    log.debug(json.dumps(dict3))
                elif len(contracts) == 0:
                    cntrNum = "None"
                    log.warning("/%(schm)s/%(tmpl)s template does not have any contract created" % dict(
                        schm=schmDisplayName, tmpl=tmplName))
                dict3 = {'Contract Name': cntrName, 'Contract Display Name': cntrDisplayName,
                         'ContractContract Number': cntrNum, 'Contract UUID': cntrUUID, 'Contract Reference': cntrRef}
                log.debug(json.dumps(dict3))
                cntr_dict = {**dict1, **dict2, **dict3}
                data.append(cntr_dict)
                log.info("Normalized /%(schm)s/%(tmpl)s/%(cntr)s contract template data" % dict(
                    schm=schmDisplayName, tmpl=tmplName, cntr=cntrName
                ))
    return data


def normalize_contract_site_data(schemas_data: dict):
    data = []
    for schema in schemas_data['schemas']:
        schmId = schema['id']
        schmDisplayName = schema['displayName']
        dict1 = {'Schema ID': schmId}
        if len(schema['sites']) >= 1:
            for siteNum, site in enumerate(schema['sites']):
                siteId = site['siteId']
                siteTmplName = site['templateName']
                siteTmplID = site['templateID']
                dict2 = {'Site ID': siteId, 'Site Template ID': siteTmplID}
                log.debug(json.dumps({**dict1, **dict2}))
                contracts = site['contracts']
                for cntrNum, cntr in enumerate(contracts):
                    cntrRef, dict3 = ("None",) * 2
                    if len(contracts) >= 1:
                        cntrRef = cntr['contractRef']
                    elif len(contracts) == 0:
                        cntrNum = "None"
                        log.warning(
                            "/%(schm)s/%(tmpl)s template does not have contracts at the site (%(site)s) level" % dict(
                                schm=schmDisplayName, tmpl=siteTmplName, site=siteId))
                    dict3 = {'Contract Number': cntrNum, 'Contract Reference': cntrRef}
                    log.debug(json.dumps(dict3))
                    cntr_dict = {**dict1, **dict2, **dict3}
                    data.append(cntr_dict)
                    log.info("Normalized /%(schm)s/%(tmpl)s/%(cntr)s contract site data" % dict(
                        schm=schmDisplayName, tmpl=siteTmplName, cntr=cntrRef
                    ))
        elif len(schema['sites']) == 0:
            log.warning("%s schema is not currently associated to any site" % schmDisplayName)
    return data


def main(*args, **kwargs):
    global src_epg_site_data
    if parser.get:
        filename = parser.filename
        session = get_ndo_session(**kwargs)
        ndo_sites = get_ndo_sites(session, **kwargs)
        ndo_sites_json = ndo_sites.json()
        site_data = normalize_sites_data(ndo_sites_json)
        write_to_excel(filename, "Infra", site_data, startCol=1, startRow=3)
        ndo_tenants = get_ndo_tenants(session, **kwargs)
        ndo_tenants_json = ndo_tenants.json()
        tenants_data = normalize_tenants_data(ndo_tenants_json)
        write_to_excel(filename, "Infra", tenants_data, startCol=7, startRow=3)
        ndo_schemas = get_ndo_schemas(session, **kwargs)
        ndo_schemas_json = ndo_schemas.json()
        schema_template_data = normalize_schema_template_data(ndo_schemas_json)
        write_to_excel(filename, "Schemas", schema_template_data, startCol=1, startRow=3)
        schema_site_data = normalize_schema_site_data(ndo_schemas_json)
        write_to_excel(filename, "Schemas", schema_site_data, startCol=13, startRow=3)
        vrf_template_data = normalize_vrf_template_data(ndo_schemas_json)
        write_to_excel(filename, "VRF Data", vrf_template_data, startCol=1, startRow=3)
        vrf_site_data = normalize_vrf_site_data(ndo_schemas_json)
        write_to_excel(filename, "VRF Data", vrf_site_data, startCol=11, startRow=3)
        bd_template_data = normalize_bd_template_data(ndo_schemas_json)
        write_to_excel(filename, "BD Data", bd_template_data, startCol=1, startRow=3)
        # Need to add L3Out data in the BD Site Json
        bd_site_data = normalize_bd_site_data(ndo_schemas_json)
        write_to_excel(filename, "BD Data", bd_site_data, startCol=14, startRow=3)
        epg_template_data = normalize_epg_template_data(ndo_schemas_json)
        write_to_excel(filename, "EPG Data", epg_template_data, startCol=1, startRow=3)
        epg_site_data = normalize_epg_site_data(ndo_schemas_json)
        write_to_excel(filename, "EPG Data", epg_site_data, startCol=17, startRow=3)
        contract_template_data = normalize_contract_template_data(ndo_schemas_json)
        write_to_excel(filename, "Contract Data", contract_template_data, startCol=1, startRow=3)
        contract_site_data = normalize_contract_site_data(ndo_schemas_json)
        write_to_excel(filename, "Contract Data", contract_site_data, startCol=11, startRow=3)

    elif parser.put:
        filename = parser.filename
        sheetname = "EPG Selection"
        # Try to load the workbook with name 'filename'
        # If the file doesn't exist or doesn't contain a sheet named 'EPG Selection', the program will end.
        try:
            workbook = load_workbook(filename, data_only=True)
            sheet = workbook[sheetname]
        except FileNotFoundError as e:
            log.error("%s file not found in the local directory. If necessary run the script with the -g option "
                      "to generate a migration file" % filename)
            sys.exit()
        except KeyError as e:
            log.error("%s sheet does not exist in %s" % (sheetname, filename))
            sys.exit()

        # Templates for the NDO objects that will be used
        ndo_bd_tmpl_dict = {
            "arpFlood": None, "description": "", "dhcpLabels": [], "displayName": "", "intersiteBumTrafficAllow": None,
            "l2Stretch": None, "l2UnknownUnicast": "", "l3MCast": None, "multiDstPktAct": "", "name": "",
            "optimizeWanBandwidth": None, "subnets": [], "unicastRouting": None, "unkMcastAct": "",
            "v6unkMcastAct": "","vmac": "", "vrfRef": ""
        }
        ndo_bd_site_dict = {
            "bdRef": "", "hostBasedRouting": None, "l3OutRefs": [], "l3Outs": [], "mac": "", "subnets": []
        }
        ndo_epg_tmpl_dict = {
            "bdRef": "", "contractRelationships": [], "description": "", "displayName": "", "epgType": "",
            "intraEpg": "", "mCastSource": None, "name": "", "preferredGroup": None, "proxyArp": None, "selectors": [],
            "subnets": [], "uSegAttrs": [], "uSegEpg": None
        }
        ndo_epg_site_dict = {
            "domainAssociations": [], "epgRef": "", "selectors": [], "staticLeafs": [], "staticPorts": [],
            "subnets": [], "uSegAttrs": []
        }

        # Keys to replace the column names in the Excel file to make them look like the keys used by the NDO objects
        src_data_keys = [
            "siteName", "siteId", "tenantName", "tenantId", "schemaName", "schemaId", "templName", "templId",
            "templVal", "siteVal", "vrfRef", "vrfName", "bdName", "bdUUID", "bdRef", "bdTemplId", "bdL2Stretch",
            "bdTemplSubnet", "bdSiteId", "bdSiteSubnet", "bdVrfRef", "anpName", "anpUUID", "anpRef", "anpTemplId",
            "anpSiteId", "epgName", "epgUUID", "epgRef", "epgTemplId", "epgSiteId"
        ]
        dst_data_keys = [
            "siteName", "siteId", "tenantName", "tenantId", "schemaName", "schemaId", "templName", "templId",
            "templVal", "siteVal", "vrfRef", "vrfName", "bdName", "bdL2Stretch", "bdL3Out1", "bdL3OutRef1", "bdL3Out2",
            "bdL3OutRef2", "anpName", "anpRef", "anpId", "epgName", "consContract"
        ]

        # Variable definition
        src_epgs = []
        dst_epgs = []
        src_bd_tmpl_data = {}
        src_bd_site_data = {}
        src_epg_tmpl_data = {}
        src_epg_site_data = {}

        # Get the Source EPG column names
        columns = [key for key in src_data_keys]

        # Iterate over each row in the sheet starting from row 303 and finishing on row 343
        for row in sheet.iter_rows(min_row=303, max_row=343, max_col=31, values_only=True):
            # Create a dictionary to store the row data
            src_data = dict(zip(columns, row))
            if src_data['siteName'] is not None:
                src_epgs.append(src_data)

        # Get the Destination EPG column names
        columns = [key for key in dst_data_keys]

        # Iterate over each row in the sheet starting from row 349 and finishing on row 389
        for row in sheet.iter_rows(min_row=349, max_row=389, max_col=23, values_only=True):
            # Create a dictionary to store the row data
            dst_data = dict(zip(columns, row))
            if dst_data['siteName'] is not None:
                dst_epgs.append(dst_data)

        # Get an NDO session token
        session = get_ndo_session(**kwargs)
        # Get all the schemas information
        ndo_schemas = get_ndo_schemas(session, **kwargs)
        ndo_schemas_json = ndo_schemas.json()

        # Verify that the amount of src and dst EPGs are the same
        # Quit the program if the values don't match
        try:
            assert len(src_epgs) == len(dst_epgs)
            log.info("Length of source and destination EPG data matches")
        except AssertionError as e:
            log.error("Source and destination EPG quantities don't match...Exiting")
            sys.exit()

        items = len(src_epgs)
        for i in range(items):
            try:
                # Iterate through the schemas to get configuration parameters for dst tenant objects
                for schema in ndo_schemas_json['schemas']:
                    if schema['id'] == src_epgs[i]['schemaId']:
                        for tmpl in schema['templates']:
                            if tmpl['templateID'] == src_epgs[i]['templId']:
                                bd_tmpl_id = src_epgs[i]['bdTemplId']
                                src_bd_tmpl_data = tmpl['bds'][bd_tmpl_id]

                                anp_tmpl_id = src_epgs[i]['anpTemplId']
                                epg_tmpl_id = src_epgs[i]['epgTemplId']
                                src_epg_tmpl_data = tmpl['anps'][anp_tmpl_id]['epgs'][epg_tmpl_id]

                        for site in schema['sites']:
                            if site['siteId'] == src_epgs[i]['siteId']:
                                bd_site_id = src_epgs[i]['bdSiteId']
                                src_bd_site_data = site['bds'][bd_site_id]

                                anp_site_id = src_epgs[i]['anpSiteId']
                                epg_site_id = src_epgs[i]['epgSiteId']
                                src_epg_site_data = site['anps'][anp_site_id]['epgs'][epg_site_id]

                # Create dictionaries based on the templates
                dst_bd_tmpl_data: dict = ndo_bd_tmpl_dict
                dst_bd_site_data: dict = ndo_bd_site_dict

                # Create dictionaries for destination BD template and site data:
                # If the src BD is NOT stretched
                if src_bd_tmpl_data['l2Stretch'] is True:
                    try:
                        vmac = src_bd_tmpl_data['vmac']
                    except KeyError as e:
                        log.debug("Src BD data does not have a vMAC configured")
                        src_bd_tmpl_data.setdefault("vmac", "")
                    for key, value in dst_bd_tmpl_data.items():
                        dst_bd_tmpl_data[key] = src_bd_tmpl_data[key]

                    for key, value in dst_bd_site_data.items():
                        dst_bd_site_data[key] = src_bd_site_data[key]

                # If the src BD is stretched
                elif src_bd_tmpl_data['l2Stretch'] is False:
                    try:
                        vmac = src_bd_tmpl_data['vmac']
                    except KeyError as e:
                        log.debug("Src BD data does not have a vMAC configured")
                        src_bd_tmpl_data.setdefault("vmac", "")
                    for key, value in dst_bd_tmpl_data.items():
                        dst_bd_tmpl_data[key] = src_bd_tmpl_data[key]
                    dst_bd_tmpl_data['arpFlood'] = True
                    dst_bd_tmpl_data['intersiteBumTrafficAllow'] = True
                    dst_bd_tmpl_data['l2Stretch'] = True
                    dst_bd_tmpl_data['optimizeWanBandwidth'] = True
                    dst_bd_tmpl_data['subnets'] = src_bd_site_data['subnets']

                    for key, value in dst_bd_site_data.items():
                        dst_bd_site_data[key] = src_bd_site_data[key]

                l3outs = [dst_epgs[i]['bdL3Out1'], dst_epgs[i]['bdL3Out2']]
                l3out_refs = [dst_epgs[i]['bdL3OutRef1'], dst_epgs[i]['bdL3OutRef2']]
                dst_bd_site_data['l3Outs'] = l3outs
                dst_bd_site_data['l3OutRefs'] = l3out_refs
                dst_bd_tmpl_data['name'] = dst_epgs[i]['bdName']
                dst_bd_tmpl_data['displayName'] = dst_epgs[i]['bdName']
                dst_bd_tmpl_data['vrfRef'] = dst_epgs[i]['vrfRef']

                for subnet in dst_bd_tmpl_data['subnets']:
                    if subnet['ip'].startswith("10.14") and dst_epgs[i]['bdName'].startswith('WB'):
                        dst_bd_site_data['hostBasedRouting'] = True
                    elif subnet['ip'].startswith("10.17") and dst_epgs[i]['bdName'].startswith('YK'):
                        dst_bd_site_data['hostBasedRouting'] = True
                    else:
                        dst_bd_site_data['hostBasedRouting'] = False

                dst_bd_ref = "/schemas/%(schm)s/templates/%(tmpl)s/bds/%(bd)s" % dict(
                    schm=dst_epgs[i]['schemaId'], tmpl=dst_epgs[i]['templName'], bd=dst_epgs[i]['bdName'])
                dst_bd_site_data['bdRef'] = dst_bd_ref

                # Create dictionaries for destination EPG template and site data:
                dst_epg_tmpl_data: dict = ndo_epg_tmpl_dict
                dst_epg_site_data: dict = ndo_epg_site_dict
                for key, value in dst_epg_tmpl_data.items():
                    dst_epg_tmpl_data[key] = src_epg_tmpl_data[key]

                dst_epg_tmpl_data['bdRef'] = dst_bd_site_data['bdRef']
                dst_epg_tmpl_data['name'] = dst_epgs[i]['epgName']
                dst_epg_tmpl_data['displayName'] = dst_epgs[i]['epgName']
                dst_epg_tmpl_data['preferredGroup'] = True
                contract_relationships = []
                contract = {"contractRef": dst_epgs[i]['consContract'], "relationshipType": "consumer"}
                contract_relationships.append(contract)

                for key, value in dst_epg_site_data.items():
                    dst_epg_site_data[key] = src_epg_site_data[key]
                dst_epg_site_data['epgRef'] = "%(anpR)s/epgs/%(epg)s" % dict(
                    anpR=dst_epgs[i]['anpRef'], epg=dst_epg_tmpl_data['name'])

                site_id = dst_epgs[i]['siteId']
                schm_id = dst_epgs[i]['schemaId']
                tmpl_name = dst_epgs[i]['templName']

                bd_ref = dst_bd_site_data['bdRef']
                epg_ref = dst_epg_site_data['epgRef']

                # Create new BD on dst tenant template
                patch_ndo_tmpl_bds(session, dst_bd_tmpl_data, bdRef=bd_ref, oper="add", **kwargs)
                # Change configuration parameters of new BD on dst tenant site
                patch_ndo_site_bds(session, dst_bd_site_data, siteId=site_id, oper="replace", **kwargs)
                # Deploy the template to the tenant for changes to take effect on APIC
                deploy = deploy_ndo_template(session, schm=schm_id, tmpl=tmpl_name)
                # Gets the deployment taskId for completeness verification
                deploy_id = deploy.json()['id']
                # Verification that the new BD has been successfully created in the dst tenant
                ndo_deploy_status_check(session, id=deploy_id)

                # Create the new EPG on the dst tenant template
                patch_ndo_tmpl_epgs(session, dst_epg_tmpl_data, epgRef=epg_ref, oper="add", **kwargs)
                # Change configuration parameters of new EPG on dst tenant site
                patch_ndo_site_epgs(session, dst_epg_site_data, siteId=site_id, oper="replace", **kwargs)

                # Information of the BD/EPG that is going to be removed from the src tenant
                src_schm_id = src_epgs[i]['schemaId']
                src_tmpl_name = src_epgs[i]['templName']
                src_site_id = src_epgs[i]['siteId']
                src_bf_ref = src_bd_site_data['bdRef']
                src_epg_ref = src_epg_site_data['epgRef']
                # The empty list is to remove the ports from the src EPG before deploying them on the dst tenant
                empty_ports = []

                # Replace operation on the src tenant EPG to remove the ports from the template
                patch_ndo_epg_static_ports(
                    session, src_epg_site_data, siteId=src_site_id, ports=empty_ports, oper="replace", **kwargs)
                # Deploy the changes to the src tenant EPG
                # At this point the only change is removing the static ports
                deploy = deploy_ndo_template(session, schm=src_schm_id, tmpl=src_tmpl_name, **kwargs)
                # Gets the deployment taskId for completeness verification
                deploy_id = deploy.json()['id']
                # Verification that the src EPG has been reconfigured
                ndo_deploy_status_check(session, id=deploy_id)

                # We allow time for ACI switches to completely remove old EPG data before applying the new one
                grace_time = 2  # Grace time in secs
                for t in trange(0, grace_time * 100, ncols=100,
                                desc="%s sec grace time for leafs to reconfigure: " % grace_time):
                    sleep(0.01)

                # Deploy the changes on the dst tenant template
                # This will create the new EPG on the dst tenant with all the imported information from the src EPG
                deploy = deploy_ndo_template(session, schm=schm_id, tmpl=tmpl_name)
                # Gets the deployment taskId for completeness verification
                deploy_id = deploy.json()['id']
                # Verification that the dst EPG has been created on the
                ndo_deploy_status_check(session, id=deploy_id)

                # Remove the src EPG from the src tenant template
                patch_ndo_tmpl_epgs(session, src_epg_tmpl_data, epgRef=src_epg_ref, oper="remove", **kwargs)
                # Remove the src BD from the src tenant template
                patch_ndo_tmpl_bds(session, src_bd_tmpl_data, bdRef=src_bf_ref, oper="remove", **kwargs)

                # Deploy the changes to the template to remove the BD/EPG from the src tenant
                deploy = deploy_ndo_template(session, schm=src_schm_id, tmpl=src_tmpl_name, **kwargs)
                # Gets the deployment taskId for completeness verification
                deploy_id = deploy.json()['id']
                # Verification that the src BD/EPG has been removed successfully
                ndo_deploy_status_check(session, id=deploy_id)

                log.info("[%(i)s/%(t)s] %(name)s BD/EPG migration completed" % dict(
                    i=(i + 1), t=items, name=dst_epg_tmpl_data['name']))
                if i < (items - 1):
                    input("Please press enter to continue to the next BD/EPG: ")

            except IndexError as e:
                log.warning("Source BD/EPG information for %(epg)s is not valid...skipping" % dict(
                    sht=sheetname, epg=src_epgs[i]['epgName']))

        log.info("All rows in the migration file have been processed!")
        sleep(1)


if __name__ == "__main__":
    main(verify=parser.ssl)
